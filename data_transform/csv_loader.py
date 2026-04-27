import glob
import os 
from data_transform.decorators import error_handler
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
import pandas as pd



class CSVLoader:
    def __init__(self, variables):
        self.variables = variables
    
    def load_csv(self, path):
        csv_files = glob.glob(os.path.join(path, '*.csv'))        

        if not csv_files:
            raise FileNotFoundError('CSV files missing.')
        
        df = pd.concat(
            [pd.read_csv(p,
                         usecols=self.variables.NEEDED_COLUMNS,
                         dtype=self.variables.COLUMNS_TYPES,
                         parse_dates=self.variables.DATE_COLUMNS) for p in csv_files],
            axis=0, 
            ignore_index=True)

        df = df.loc[~df['AssociationId'].isin(self.variables.INVALID_ASSOCIATION_ID)]
        
        df.drop_duplicates(subset=self.variables.REMOVE_DUPLICATES, inplace=True)
        
        return df

    def export(self, df, path, file_name):
        file = os.path.join(path, file_name)
        self._export_dataframe_to_excel(df, file)
        self._create_name_definition(file)

        
    @error_handler
    def _export_dataframe_to_excel(self, df, file_name):
        df.to_excel(file_name,
            sheet_name=self.variables.SHEET_NAME, 
            index=False, 
            startrow=self.variables.START_ROW)

    @error_handler
    def _create_name_definition(self, file_name):
        wb = load_workbook(file_name)
        ws = wb[self.variables.SHEET_NAME]

        last_row = ws.max_row
        last_col = ws.max_column

        ref = f"{self.variables.SHEET_NAME}!$A${self.variables.START_ROW + 1}:${chr(64 + last_col)}${last_row}"

        named_range = DefinedName(name=self.variables.DEFINED_NAME, attr_text=ref)
        wb.defined_names.add(named_range)
        
        wb.save(file_name)